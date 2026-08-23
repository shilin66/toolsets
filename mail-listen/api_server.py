#!/usr/bin/env python
"""
邮件监听系统 API 服务
"""
from flask import Flask, request, jsonify, send_file, url_for
from datetime import datetime
from zoneinfo import ZoneInfo
from loguru import logger
import sys
import json
import re
import uuid
from functools import wraps
from copy import copy
from io import BytesIO
from pathlib import Path

import requests

from pydantic import ValidationError

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None

from database import email_db
from actions import build_security_time, extract_fastgpt_reply
from config import settings, log_format
from email_client import EmailClient
from mail_accounts import (
    MailAccountConfig,
    account_row_to_config,
    encrypt_password,
    list_enabled_account_configs,
)
from mail_listener import listener_manager
from email_attachments import (
    attachment_path_from_relative,
    build_attachment_url,
    save_preview_attachment,
)
from html_image import image_path_from_filename
from models import EmailMessage, APIRequest
from cutover_prompt import build_cutover_extract_prompt, get_fixed_field_definitions
from cutover_report import CutoverReportNotReady, report_cutover_task
from cutover_task import (
    CUTOVER_SCENE_NORMAL,
    REPLY_STATUS_PENDING,
    VALID_CUTOVER_SCENES,
    VALID_CUTOVER_TAGS,
    WRITABLE_CUTOVER_SCENES,
    apply_task_edit,
    confirm_task,
    cutover_line_type_label,
    cutover_scene_label,
    cutover_task_status_label,
    get_cutover_task_detail,
    save_cutover_fill_tasks,
    switch_task_line_type,
)
from record_query import RecordQueryRepository, build_query_page
from supplier_circuits import (
    build_circuits_workbook_bytes,
    migrate_supplier_circuits_from_xlsx,
    parse_circuits_workbook,
)
from supplier_config import (
    MAIL_TYPE_OPTIONS,
    SupplierConfigConflictError,
    SupplierConfigCreate,
    SupplierConfigRepository,
    SupplierConfigUpdate,
    build_mail_classify_prompt,
)

app = Flask(__name__)

# 配置日志（统一格式：带邮箱标识，见 config.log_format）
logger.remove()
logger.add(
    sys.stdout,
    level="INFO",
    format=log_format,
)

# API Key 配置
API_KEY = settings.api_key
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
TEMPLATE_XLSX_PATH = BASE_DIR / 'template.xlsx'
TEMPLATE_EXPORT_SHEETS = ('电路表', '割接原因表')
DEFAULT_LINE_QUERY_KEYWORDS = ('CircuitID', 'OrderNumber', 'InternationalId', 'CircuitIDRT')

def require_api_key(f):
    """API Key 校验装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # 获取 Authorization 头
        auth_header = request.headers.get('Authorization')
        
        if not auth_header:
            return jsonify({
                'success': False,
                'message': '缺少 Authorization 头'
            }), 401
        
        # 检查 Bearer 格式
        if not auth_header.startswith('Bearer '):
            return jsonify({
                'success': False,
                'message': 'Authorization 头格式错误，应为: Bearer <api_key>'
            }), 401
        
        # 提取 API Key
        provided_key = auth_header[7:]  # 去掉 "Bearer " 前缀
        
        if provided_key != API_KEY:
            logger.warning("API Key 校验失败")
            return jsonify({
                'success': False,
                'message': 'API Key 无效'
            }), 401
        
        return f(*args, **kwargs)
    
    return decorated_function


def get_supplier_config_repository() -> SupplierConfigRepository:
    return SupplierConfigRepository(email_db)


def get_record_query_repository() -> RecordQueryRepository:
    return RecordQueryRepository(email_db)


def validation_error_message(error: ValidationError) -> str:
    fields = []
    for item in error.errors():
        loc = item.get('loc') or ()
        if loc:
            fields.append('.'.join(str(part) for part in loc))

    if fields:
        return f"参数校验失败：{', '.join(fields)}"
    return "参数校验失败"

def parse_time(time_str, field_name="time"):
    """
    解析时间字符串，支持多种格式
    
    Args:
        time_str: 时间字符串
        field_name: 字段名称（用于日志）
    
    Returns:
        datetime: 解析后的时间对象
    """
    if not time_str:
        return datetime.now()
    
    try:
        # 首选格式：2025-11-01 22:53:17
        return datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        try:
            # 备选格式：ISO 格式
            return datetime.fromisoformat(time_str.replace('Z', '+00:00'))
        except ValueError:
            logger.warning(f"{field_name} 格式错误，使用当前时间: {time_str}")
            return datetime.now()


def parse_required_time(time_str, field_name="time"):
    """解析必填时间字段，格式错误时返回错误信息。"""
    if not time_str:
        return None, f'缺少必需参数：{field_name}'

    try:
        return datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S'), None
    except ValueError:
        try:
            return datetime.fromisoformat(str(time_str).replace('Z', '+00:00')), None
        except ValueError:
            return None, f'{field_name} 格式错误，应为 YYYY-MM-DD HH:MM:SS 或 ISO 格式'


def snapshot_row_style(ws, row):
    """缓存模板数据行样式，避免清空数据后丢失格式。"""
    row_styles = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        row_styles.append({
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format,
            'protection': copy(cell.protection),
            'has_style': cell.has_style,
        })

    return {
        'height': ws.row_dimensions[row].height,
        'cells': row_styles,
    }


def apply_row_style(ws, target_row, style_snapshot):
    for col, cell_style in enumerate(style_snapshot['cells'], start=1):
        if not cell_style['has_style']:
            continue

        target_cell = ws.cell(row=target_row, column=col)
        target_cell.font = copy(cell_style['font'])
        target_cell.fill = copy(cell_style['fill'])
        target_cell.border = copy(cell_style['border'])
        target_cell.alignment = copy(cell_style['alignment'])
        target_cell.number_format = cell_style['number_format']
        target_cell.protection = copy(cell_style['protection'])

    if style_snapshot['height']:
        ws.row_dimensions[target_row].height = style_snapshot['height']


def normalize_sheet_rows(rows, headers, sheet_name):
    """
    将 API 入参转换为按模板表头排序的二维行数据。

    支持两种格式：
    1. [{"表头1": "值1"}]
    2. [["值1", "值2"]]
    """
    if rows is None:
        return None
    if not isinstance(rows, list):
        raise ValueError(f'{sheet_name} 数据必须是数组')

    normalized_rows = []
    for index, row in enumerate(rows, start=1):
        if isinstance(row, dict):
            normalized_rows.append([row.get(header) for header in headers])
        elif isinstance(row, list):
            normalized_rows.append(row[:len(headers)] + [None] * max(len(headers) - len(row), 0))
        else:
            raise ValueError(f'{sheet_name} 第 {index} 行必须是对象或数组')

    return normalized_rows


def replace_sheet_data(ws, rows):
    data_row_style = snapshot_row_style(ws, 2)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_index, row_values in enumerate(rows, start=2):
        apply_row_style(ws, row_index, data_row_style)
        for col_index, value in enumerate(row_values, start=1):
            ws.cell(row=row_index, column=col_index, value=value)


def normalize_xlsx_filename(filename):
    filename = Path(str(filename or 'template.xlsx')).name.strip()
    if not filename:
        filename = 'template.xlsx'
    if not filename.endswith('.xlsx'):
        filename = f'{filename}.xlsx'
    return filename


def build_cutover_filename(supplier, carrier_ticket_no):
    safe_supplier = normalize_excel_value(supplier) or 'cutover'
    safe_ticket_no = normalize_excel_value(carrier_ticket_no).replace('/', '_').replace('\\', '_')
    return f'{safe_supplier}{safe_ticket_no}.xlsx'


def save_workbook_output(output, filename):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    output_path = DATA_DIR / normalize_xlsx_filename(filename)
    output_path.write_bytes(output.getvalue())
    return output_path


def build_template_workbook(payload=None):
    if load_workbook is None:
        raise RuntimeError('缺少依赖 openpyxl，无法生成 Excel 文件')
    if not TEMPLATE_XLSX_PATH.exists():
        raise FileNotFoundError(f'模板文件不存在: {TEMPLATE_XLSX_PATH}')

    wb = load_workbook(TEMPLATE_XLSX_PATH)

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in TEMPLATE_EXPORT_SHEETS:
            del wb[sheet_name]

    if payload:
        sheet_payload_keys = {
            '电路表': ('circuits', 'circuit_rows', '电路表'),
            '割接原因表': ('reasons', 'reason_rows', '割接原因表'),
        }

        for sheet_name, keys in sheet_payload_keys.items():
            ws = wb[sheet_name]
            headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            rows = next((payload.get(key) for key in keys if key in payload), None)
            normalized_rows = normalize_sheet_rows(rows, headers, sheet_name)
            if normalized_rows is not None:
                replace_sheet_data(ws, normalized_rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def normalize_excel_value(value):
    if value is None:
        return ''
    return str(value).strip()


@app.route('/api/cutover/template-columns', methods=['GET'])
@require_api_key
def get_cutover_template_columns():
    """返回客户线路 Excel 模板各工作表的表头顺序，供前端按导出顺序展示。"""
    try:
        if load_workbook is None:
            raise RuntimeError('缺少依赖 openpyxl，无法读取模板文件')
        if not TEMPLATE_XLSX_PATH.exists():
            raise FileNotFoundError(f'模板文件不存在: {TEMPLATE_XLSX_PATH}')

        wb = load_workbook(TEMPLATE_XLSX_PATH, read_only=True, data_only=True)
        try:
            columns = {}
            for sheet_name in TEMPLATE_EXPORT_SHEETS:
                ws = wb[sheet_name]
                columns[sheet_name] = [
                    normalize_excel_value(ws.cell(row=1, column=col).value)
                    for col in range(1, ws.max_column + 1)
                    if normalize_excel_value(ws.cell(row=1, column=col).value)
                ]
        finally:
            wb.close()

        return jsonify({
            'success': True,
            'message': '模板表头查询成功',
            'data': {
                'circuit': columns.get('电路表', []),
                'reason': columns.get('割接原因表', []),
            }
        }), 200
    except Exception as e:
        logger.error(f"查询割接模板表头失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


def query_supplier_circuits(supplier, keywords):
    """按供应商与关键字查询线路表（数据库），关键字子串匹配 Supplier Circuit ID。

    Circuit ID 单元格可能含换行分隔的多条线路，此处保持整格存储，不做拆分。
    """
    supplier = normalize_excel_value(supplier)
    normalized_keywords = []
    for keyword in keywords:
        normalized_keyword = normalize_excel_value(keyword)
        if normalized_keyword:
            normalized_keywords.append(normalized_keyword)

    if not normalized_keywords:
        return []

    results = []
    seen = set()
    for row_data in email_db.get_all_supplier_circuits():
        row_supplier = normalize_excel_value(row_data.get('supplier'))
        if supplier and row_supplier != supplier:
            continue

        supplier_circuit_id = normalize_excel_value(row_data.get('supplier_circuit_id'))
        if not any(keyword in supplier_circuit_id for keyword in normalized_keywords):
            continue

        response_row = {
            'supplier': row_supplier,
            'supplier_circuit_id': supplier_circuit_id,
            'circuit_id': normalize_excel_value(row_data.get('circuit_id')),
            'line_type': normalize_excel_value(row_data.get('line_type')),
            'remark': normalize_excel_value(row_data.get('remark')),
        }
        dedupe_key = tuple(response_row.values())
        if dedupe_key in seen:
            continue

        seen.add(dedupe_key)
        results.append(response_row)

    return results


def match_supplier_circuits_by_payload(payload):
    """按割接填报入参逐条线路重新查询本地线路表，返回每条线路的匹配结果。"""
    payload = payload or {}
    supplier = payload.get('supplier')
    line_array = payload.get('line_array')
    if line_array is None:
        line_array = (payload.get('line_array_info') or {}).get('data') or []
    line_query_keywords = payload.get('line_query_keywords') or list(DEFAULT_LINE_QUERY_KEYWORDS)
    if not isinstance(line_array, list):
        raise ValueError('line_array 必须是数组')

    lines = []
    for line in line_array:
        if not isinstance(line, dict):
            continue
        keywords = [keyword for keyword in line_keywords(line, line_query_keywords) if keyword]
        lines.append({
            'line': line,
            'keywords': keywords,
            'circuits': query_supplier_circuits(supplier, keywords),
        })

    return {
        'supplier': normalize_excel_value(supplier),
        'lines': lines,
    }


def parse_cutover_datetime(value, timezone_name='UTC'):
    value = normalize_excel_value(value)
    if not value:
        return None

    formats = (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%d %b %Y %H:%M',
        '%d %B %Y %H:%M',
    )
    for date_format in formats:
        try:
            parsed = datetime.strptime(value, date_format)
            return parsed.replace(tzinfo=ZoneInfo(timezone_name or 'UTC'))
        except ValueError:
            continue

    raise ValueError(f'时间格式错误: {value}')


def format_beijing_time(value, timezone_name='UTC'):
    parsed = parse_cutover_datetime(value, timezone_name)
    if parsed is None:
        return ''
    return parsed.astimezone(ZoneInfo('Asia/Shanghai')).strftime('%Y-%m-%d %H:%M:%S')


def split_time_range(time_range, separator='/'):
    parts = [part.strip() for part in normalize_excel_value(time_range).split(separator, 1)]
    if len(parts) != 2 or not parts[0] or not parts[1]:
        raise ValueError(f'时间范围格式错误: {time_range}')
    return parts


def split_impact_time_range(impact_datetime):
    value = normalize_excel_value(impact_datetime).replace('(UTC)', '').strip()
    if '/' in value:
        return split_time_range(value, '/')
    if ' - ' in value:
        return split_time_range(value, ' - ')
    raise ValueError(f'ImpactDateTime 格式错误: {impact_datetime}')


def calculate_interrupt_duration(impact_duration, interruptions_counts):
    duration_text = normalize_excel_value(impact_duration)
    try:
        count = int(float(interruptions_counts or 1))
    except (TypeError, ValueError):
        count = 1

    duration_number = 0
    digits = ''
    for char in duration_text:
        if char.isdigit():
            digits += char
        elif digits:
            break

    if digits:
        duration_number = int(digits)

    minutes = duration_number * count
    if 'h' in duration_text.lower():
        minutes = duration_number * 60 * count

    if minutes < 5:
        return '瞬断'
    return f'中断：{minutes}分钟'


def build_customer_circuit(line, circuit_info, line_name, cutover_time, cutover_timezone):
    return {
        '客户名称': None,
        '电路代号': circuit_info.get('circuit_id'),
        '客户电路在割接段是否有保护': '无保护',
        '预计影响客户业务时长': calculate_interrupt_duration(
            line.get('ImpactDuration'),
            line.get('InteruptionsCounts'),
        ),
        '割接线路/设备名称': line_name,
        '割接时间': cutover_time,
        '割接省': '国际公司',
        '时间类型': cutover_timezone,
        '客户编码': None
    }


def build_customer_reason(line_name, cutover_time, cutover_timezone, cutover_reason, location):
    return {
        '割接线路/设备名称': line_name,
        '割接原因': f'{line_name}--Work Location:{location}. Reason: {cutover_reason}',
        '割接省': '国际公司',
        '中继段': None,
        '割接时间': cutover_time,
        '割接分类': '线路割接',
        '割接原因分类': '其它运营商割接',
        '割接类型': '紧急割接',
        '割接级别': '其它运营商',
        '时间类型': cutover_timezone
    }


def backbone_fixed_sections(title, cutover_reason, location):
    """骨干线路填报中与客户线路无关的固定分组，供生成与类型切换复用。"""
    return {
        '人员信息': {
            '现场指挥人员名称': 'RT',
            '支撑人员': 'NOC',
            '省公司审核人姓名': '钟悦',
            '业务测试人员': 'NOC',
            '操作审核人姓名': 'TTK NOC',
            '现场指挥人电话': 'NA',
            '支撑人电话': '固定值',
            '省公司审核人电话': '固定值',
            '业务测试人员电话': '固定值',
            '操作审核人电话': 'NA',
            '现场指挥人单位': '国际公司',
            '支撑人单位': '国际公司',
            '省公司审核人单位': '国际公司',
            '业务测试人单位': '国际公司',
            '操作审核人单位': '国际公司'
        },
        '关联网络资源变更单': {
            '是否设计网络资源变更': '否'
        },
        '其他': {
            '影响军队': '否',
            '是否变更网管中网元管理对象配置': '否',
            '是否增加、删除或更改光开关(OLP)': '否',
            '是否影响联通': '否',
            '是否需要集团网管配置操作': '否',
            '割接原因': f'{title}Location:{location}. Reason: {cutover_reason}',
            '风险操作影响范围': ''
        }
    }


def build_backbone_circuit(line, circuit_info, title, cutover_timezone, cutover_reason, location):
    impact_start_raw, impact_end_raw = split_impact_time_range(line.get('ImpactDateTime'))
    impact_start_time = format_beijing_time(impact_start_raw, cutover_timezone)
    impact_end_time = format_beijing_time(impact_end_raw, cutover_timezone)
    interrupt_type = calculate_interrupt_duration(
        line.get('ImpactDuration'),
        line.get('InteruptionsCounts'),
    )

    return {
        '基本信息': {
            '标题': title,
            '割接分类': '线路割接',
            '割接原因分类': '其它运营商割接',
            '涉及系统（网络）': '',
            '操作厂家': '',
            '设备名称': '',
            '割接省份': '国际',
            '变更操作内容': '其他单段光缆割接',
            '变更操作等级': '四级',
            '调度方式': '',
            '中断类型': interrupt_type,
            '需要配合的省': '',
            '中断原因': f'Location:{location}. Reason: {cutover_reason}',
            '是否涉及集团维护设备或业务影响超出本省范围': '否',
            '是否有回退应急预案和舆情应对方案': '不涉及',
            '割接地点': location,
            '是否跨专业': '否'
        },
        '割接对象': {
            '割接类型': '正常割接',
            '割接对象所属机构': '国际',
            '割接对象类型': '高阶通道',
            '系统名称': circuit_info.get('circuit_id'),
            '割接开始时间': impact_start_time,
            '割接结束时间': impact_end_time,
            '割接名称': ''
        },
        **backbone_fixed_sections(title, cutover_reason, location)
    }


def line_keywords(line, query_fields=None):
    fields = query_fields or DEFAULT_LINE_QUERY_KEYWORDS
    return [
        line.get(field)
        for field in fields
    ]


def is_customer_line(line_type):
    return '客户' in normalize_excel_value(line_type)


def is_backbone_line(line_type):
    return '骨干' in normalize_excel_value(line_type)


def build_validation_message(message_type, message, line, keywords, circuit=None, circuits=None):
    validation_message = {
        'type': message_type,
        'message': message,
        'line': line,
        'keywords': [keyword for keyword in keywords if keyword],
    }
    if circuit is not None:
        validation_message['circuit'] = circuit
    if circuits is not None:
        validation_message['circuits'] = circuits
    return validation_message


def build_cutover_fill_response(data):
    supplier = data.get('supplier')
    carrier_ticket_no = data.get('carrier_ticket_no')
    cutover_time = data.get('cutover_time')
    cutover_timezone = data.get('cutover_timezone') or 'UTC'
    cutover_reason = data.get('cutover_reason')
    location = data.get('location')
    line_array = data.get('line_array')
    if line_array is None:
        line_array_info = data.get('line_array_info') or {}
        line_array = line_array_info.get('data') or []
    line_query_keywords = data.get('line_query_keywords') or list(DEFAULT_LINE_QUERY_KEYWORDS)

    if not normalize_excel_value(carrier_ticket_no):
        raise ValueError('缺少必需参数：carrier_ticket_no')
    if not normalize_excel_value(cutover_time):
        raise ValueError('缺少必需参数：cutover_time')
    if not isinstance(line_array, list):
        raise ValueError('line_array 必须是数组')
    if not isinstance(line_query_keywords, list):
        raise ValueError('line_query_keywords 必须是数组')

    filename = build_cutover_filename(supplier, carrier_ticket_no)
    title = f'(EUR){supplier}网内割接{carrier_ticket_no}'
    start_raw, end_raw = split_time_range(cutover_time)
    cut_start_time = format_beijing_time(start_raw, cutover_timezone)
    cut_end_time = format_beijing_time(end_raw, cutover_timezone)

    customer_circuits = []
    backbone_circuits = []
    matched_circuits = []
    validation_messages = []

    for line in line_array:
        keywords = line_keywords(line, line_query_keywords)
        supplier_circuits = query_supplier_circuits(supplier, keywords)
        if not supplier_circuits:
            validation_messages.append(build_validation_message(
                'unmatched_line',
                '线路表未查询到匹配线路',
                line,
                keywords,
            ))
            continue

        if len(supplier_circuits) > 1:
            validation_messages.append(build_validation_message(
                'multiple_matches',
                '查询到多条线路，请人工确认唯一线路后再生成填报字段',
                line,
                keywords,
                circuits=supplier_circuits,
            ))
            matched_circuits.extend(supplier_circuits)
            continue

        for circuit_info in supplier_circuits:
            circuit_id = normalize_excel_value(circuit_info.get('circuit_id'))
            line_type = normalize_excel_value(circuit_info.get('line_type'))
            matched_circuits.append(circuit_info)
            if not circuit_id:
                validation_messages.append(build_validation_message(
                    'empty_circuit_id',
                    '查询到的 circuit_id 为空，请维护线路表后再生成填报字段',
                    line,
                    keywords,
                    circuit=circuit_info,
                ))
                continue
            if not line_type:
                validation_messages.append(build_validation_message(
                    'empty_line_type',
                    '查询到的 line_type 为空，请维护线路表客户/骨干字段后再生成填报字段',
                    line,
                    keywords,
                    circuit=circuit_info,
                ))
                continue

            if is_customer_line(line_type):
                customer_circuits.append(build_customer_circuit(
                    line,
                    circuit_info,
                    title,
                    cutover_time,
                    cutover_timezone,
                ))
            elif is_backbone_line(line_type):
                backbone_circuits.append(build_backbone_circuit(
                    line,
                    circuit_info,
                    title,
                    cutover_timezone,
                    cutover_reason,
                    location,
                ))
            else:
                validation_messages.append(build_validation_message(
                    'unsupported_line_type',
                    f'不支持的 line_type: {line_type}',
                    line,
                    keywords,
                    circuit=circuit_info,
                ))

    reasons = []
    if customer_circuits:
        reasons.append(build_customer_reason(
            title,
            cutover_time,
            cutover_timezone,
            cutover_reason,
            location,
        ))

    return {
        'title': title,
        'filename': filename,
        'circuits': customer_circuits,
        'reasons': reasons,
        'backbone_circuits': backbone_circuits,
        'matched_circuits': matched_circuits,
        'validation_messages': validation_messages,
        'cutStartTime': cut_start_time,
        'cutEndTime': cut_end_time,
    }


def build_cutover_fill_msg(response_data, email_records_id):
    """生成填报结果摘要：只含摘要信息与任务详情链接，不展示具体填报字段。"""
    lines = ['割接填报摘要']
    title = response_data.get('title')
    if title:
        lines.append(f"标题: {title}")
    lines.append(f"割接窗口(北京时间): {response_data.get('cutStartTime')} 至 {response_data.get('cutEndTime')}")

    tasks = response_data.get('tasks') or []
    lines.append(f"生成任务: {len(tasks)} 个")

    validation_count = len(response_data.get('validation_messages') or [])
    if validation_count:
        lines.append(f"校验提示: {validation_count} 条，请在任务详情中人工核对")

    if tasks:
        # 任务详情直达：邮件详情页的任务展开行（taskId 参数自动展开对应任务）
        detail_base = url_for('dashboard_admin', _external=True).rstrip('/') \
            + f"/cutover-emails/{email_records_id}"
        lines.extend(['', '任务详情链接:'])
        for task in tasks:
            label = task.get('line_type_label') or task.get('line_type')
            lines.append(
                f"任务 #{task.get('id')}（{label}）: "
                f"{detail_base}?taskId={task.get('id')}"
            )

    return '\n'.join(lines)


def parse_receiver_list(receiver):
    if not receiver:
        return []
    if isinstance(receiver, list):
        return receiver
    if isinstance(receiver, str):
        try:
            parsed = json.loads(receiver)
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            pass
        return [receiver]
    return [str(receiver)]


def email_record_to_message(record):
    """将 email_records 数据库记录还原为回复邮件所需的 EmailMessage。"""
    return EmailMessage(
        uid=int(record.get('email_id')),
        subject=record.get('subject') or '',
        sender=record.get('sender') or '',
        recipients=parse_receiver_list(record.get('receiver')),
        content=record.get('content') or '',
        html_content=record.get('html_content') or None,
        message_id=record.get('message_id') or None,
        reply_to=record.get('reply_to') or None,
        references=record.get('references') or None,
        in_reply_to=record.get('in_reply_to') or None,
        received_date=parse_time(record.get('create_time'), 'create_time'),
        attachments=record.get('attachments') or [],
    )


# 首次启动时把 data/线路表.xlsx 迁移进数据库（仅数据库为空时执行）
migrate_supplier_circuits_from_xlsx()


# 人工确认回复时允许覆盖收件人，仅做基础格式校验（具体投递以 SMTP 服务器校验为准）
EMAIL_ADDRESS_PATTERN = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


def parse_reply_recipients(value):
    """解析确认回复的 recipients 参数（数组或逗号/分号/空白分隔字符串）。

    未传入返回 None（沿用原邮件 reply_to/sender）；格式非法抛 ValueError。
    """
    if value is None:
        return None
    if isinstance(value, str):
        parts = re.split(r'[,;，；\s]+', value)
    elif isinstance(value, (list, tuple)):
        parts = [str(item) for item in value]
    else:
        raise ValueError('recipients 必须是数组或字符串')
    recipients = [part.strip() for part in parts if part and part.strip()]
    if not recipients:
        raise ValueError('recipients 不能为空')
    for address in recipients:
        if not EMAIL_ADDRESS_PATTERN.match(address):
            raise ValueError(f'收件人邮箱格式非法：{address}')
    return recipients


@app.route('/health', methods=['GET'])
def health_check():
    """健康检查接口"""
    return jsonify({
        'status': 'ok',
        'message': '服务运行正常',
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })


@app.route('/api/cutover/reply', methods=['POST'])
@app.route('/api/email/reply', methods=['POST'])  # 旧路径兼容，FastGPT 存量工作流仍可使用
@require_api_key
def reply_cutover_email():
    """
    割接邮件回复登记（拒绝割接）：只落库为待确认草稿，不直接发送。
    需管理台人工确认后调 /api/cutover/reply/confirm 才真正发送。

    POST JSON:
        {
            "email_records_id": 123,   # 邮件记录主键（主参数）
            "email_id": 2450,          # 可选兼容：IMAP UID，未传 email_records_id 时回退使用
            "reply_content": "回复内容",
            "cutover_scene": "emergency"  # 可选：emergency/major_event/in_window
        }

    传入 cutover_scene 表示 FastGPT 判定为拒绝割接的特殊场景，
    随草稿一并暂存，人工确认发送成功后才写入邮件记录。
    """
    try:
        data = request.get_json(silent=True) or {}
        email_records_id = data.get('email_records_id')
        email_id = data.get('email_id')
        reply_content = data.get('reply_content')
        cutover_scene = data.get('cutover_scene')

        if (email_records_id is None and email_id is None) or not normalize_excel_value(reply_content):
            return jsonify({
                'success': False,
                'message': '缺少必需参数：email_records_id, reply_content'
            }), 400

        if cutover_scene is not None and cutover_scene not in VALID_CUTOVER_SCENES:
            return jsonify({
                'success': False,
                'message': f'cutover_scene 取值非法，可选值：{", ".join(sorted(VALID_CUTOVER_SCENES))}'
            }), 400

        try:
            if email_records_id is not None:
                email_records_id = int(email_records_id)
            elif email_id is not None:
                email_id = int(email_id)
        except (TypeError, ValueError):
            return jsonify({
                'success': False,
                'message': 'email_records_id / email_id 必须是整数'
            }), 400

        if email_records_id is not None:
            email_record = email_db.get_email_record_by_id(email_records_id)
            lookup_label = f'email_records_id={email_records_id}'
        else:
            email_record = email_db.get_email_record(email_id)
            lookup_label = f'email_id={email_id}'
        if not email_record:
            return jsonify({
                'success': False,
                'message': f'邮件记录不存在：{lookup_label}'
            }), 404

        record_id = email_record['id']
        if not email_db.save_pending_reply(record_id, normalize_excel_value(reply_content), cutover_scene or ''):
            return jsonify({
                'success': False,
                'message': f'待确认回复登记失败：{lookup_label}'
            }), 500

        logger.info(f"回复草稿已登记待人工确认：email_records_id={record_id}, scene={cutover_scene or 'normal'}")
        return jsonify({
            'success': True,
            'message': '回复已登记，待人工确认后发送',
            'data': {
                'email_records_id': record_id,
                'email_id': email_record.get('email_id'),
                'subject': email_record.get('subject') or '',
                'recipient': email_record.get('reply_to') or email_record.get('sender'),
                'reply_status': REPLY_STATUS_PENDING,
                'cutover_scene': cutover_scene or CUTOVER_SCENE_NORMAL,
                'cutover_scene_label': cutover_scene_label(cutover_scene or CUTOVER_SCENE_NORMAL),
            }
        }), 200

    except Exception as e:
        logger.error(f"邮件回复登记接口失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/reply/confirm', methods=['POST'])
@require_api_key
def confirm_cutover_reply():
    """
    人工确认发送待确认回复：校验草稿状态后发送，成功后写入割接场景。

    POST JSON:
        {
            "email_records_id": 123,   # 邮件记录主键（主参数）
            "email_id": 2450,          # 可选兼容：IMAP UID
            "reply_content": "...",    # 可选：人工修改后的回复内容，不传则用草稿内容
            "recipients": ["a@x.com", "b@x.com"]  # 可选：覆盖收件人，支持多个，不传沿用原回复地址
        }
    """
    try:
        data = request.get_json(silent=True) or {}
        email_records_id = data.get('email_records_id')
        email_id = data.get('email_id')
        reply_content = normalize_excel_value(data.get('reply_content'))

        try:
            recipients = parse_reply_recipients(data.get('recipients'))
        except ValueError as error:
            return jsonify({
                'success': False,
                'message': str(error)
            }), 400

        if email_records_id is None and email_id is None:
            return jsonify({
                'success': False,
                'message': '缺少必需参数：email_records_id'
            }), 400

        try:
            if email_records_id is not None:
                email_records_id = int(email_records_id)
            elif email_id is not None:
                email_id = int(email_id)
        except (TypeError, ValueError):
            return jsonify({
                'success': False,
                'message': 'email_records_id / email_id 必须是整数'
            }), 400

        if email_records_id is not None:
            email_record = email_db.get_email_record_by_id(email_records_id)
            lookup_label = f'email_records_id={email_records_id}'
        else:
            email_record = email_db.get_email_record(email_id)
            lookup_label = f'email_id={email_id}'
        if not email_record:
            return jsonify({
                'success': False,
                'message': f'邮件记录不存在：{lookup_label}'
            }), 404

        if (email_record.get('reply_status') or '') != REPLY_STATUS_PENDING:
            return jsonify({
                'success': False,
                'message': '该邮件当前没有待确认的回复'
            }), 400

        final_content = reply_content or (email_record.get('pending_reply_content') or '').strip()
        if not final_content:
            return jsonify({
                'success': False,
                'message': '回复内容为空，无法发送'
            }), 400

        original_email = email_record_to_message(email_record)
        reply_client = build_reply_email_client(email_record)
        if reply_client is None:
            return jsonify({
                'success': False,
                'message': '未找到可用的邮箱账号配置，无法回复'
            }), 500
        success = reply_client.reply_email(
            original_email,
            final_content,
            html_content=email_record.get('html_content') or None,
            recipients=recipients,
        )

        if not success:
            return jsonify({
                'success': False,
                'message': '邮件回复发送失败'
            }), 500

        record_id = email_record['id']
        if not email_db.mark_reply_sent(record_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')):
            logger.warning(f"回复状态标记失败：email_records_id={record_id}")

        pending_scene = email_record.get('pending_reply_scene') or ''
        if pending_scene and pending_scene != CUTOVER_SCENE_NORMAL:
            if not email_db.update_email_record_cutover_scene(record_id, pending_scene):
                logger.warning(f"邮件割接场景写入失败：email_records_id={record_id}, scene={pending_scene}")

        final_recipients = recipients or [original_email.reply_to or original_email.sender]
        return jsonify({
            'success': True,
            'message': '邮件回复发送成功',
            'data': {
                'email_records_id': record_id,
                'email_id': email_record.get('email_id'),
                'subject': original_email.subject,
                'recipient': ', '.join(final_recipients),
                'recipients': final_recipients,
                'cutover_scene': pending_scene or CUTOVER_SCENE_NORMAL,
                'cutover_scene_label': cutover_scene_label(pending_scene or CUTOVER_SCENE_NORMAL),
            }
        }), 200

    except Exception as e:
        logger.error(f"邮件回复确认接口失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/reply/cancel', methods=['POST'])
@require_api_key
def cancel_cutover_reply():
    """放弃待确认回复草稿（不发送，不写入割接场景）。"""
    try:
        data = request.get_json(silent=True) or {}
        email_records_id = data.get('email_records_id')
        email_id = data.get('email_id')

        if email_records_id is None and email_id is None:
            return jsonify({
                'success': False,
                'message': '缺少必需参数：email_records_id'
            }), 400

        try:
            if email_records_id is not None:
                email_records_id = int(email_records_id)
            elif email_id is not None:
                email_id = int(email_id)
        except (TypeError, ValueError):
            return jsonify({
                'success': False,
                'message': 'email_records_id / email_id 必须是整数'
            }), 400

        if email_records_id is not None:
            email_record = email_db.get_email_record_by_id(email_records_id)
            lookup_label = f'email_records_id={email_records_id}'
        else:
            email_record = email_db.get_email_record(email_id)
            lookup_label = f'email_id={email_id}'
        if not email_record:
            return jsonify({
                'success': False,
                'message': f'邮件记录不存在：{lookup_label}'
            }), 404

        if (email_record.get('reply_status') or '') != REPLY_STATUS_PENDING:
            return jsonify({
                'success': False,
                'message': '该邮件当前没有待确认的回复'
            }), 400

        record_id = email_record['id']
        if not email_db.cancel_pending_reply(record_id):
            return jsonify({
                'success': False,
                'message': f'取消待确认回复失败：{lookup_label}'
            }), 500

        return jsonify({
            'success': True,
            'message': '已放弃待确认回复',
            'data': {
                'email_records_id': record_id,
                'email_id': email_record.get('email_id'),
            }
        }), 200

    except Exception as e:
        logger.error(f"邮件回复取消接口失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/email/images/<filename>', methods=['GET'])
def view_email_image(filename):
    """查看 HTML 邮件正文生成的图片。"""
    try:
        image_path = image_path_from_filename(filename)
        if not image_path or not image_path.exists() or not image_path.is_file():
            return jsonify({
                'success': False,
                'message': '图片不存在'
            }), 404

        return send_file(image_path, mimetype='image/png', as_attachment=False)
    except Exception as e:
        logger.error(f"查看邮件图片失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/scene', methods=['POST'])
@require_api_key
def update_cutover_scene():
    """
    割接场景回写：FastGPT 判定特殊规则命中等场景后调用，写入邮件记录，
    供管理台列表与详情展示。命中特殊规则（rule_skipped）时不应再调用
    任务生成接口。

    POST JSON:
        {
            "email_records_id": 123,
            "cutover_scene": "rule_skipped",
            "scene_remark": "Carrier 为 China Telecom Europe Limited，命中 RT 特殊规则，不上报"
        }
    """
    try:
        data = request.get_json(silent=True) or {}
        email_records_id = data.get('email_records_id')
        cutover_scene = data.get('cutover_scene')
        scene_remark = data.get('scene_remark')

        if email_records_id is None or not cutover_scene:
            return jsonify({
                'success': False,
                'message': '缺少必需参数：email_records_id, cutover_scene'
            }), 400

        if cutover_scene not in WRITABLE_CUTOVER_SCENES:
            return jsonify({
                'success': False,
                'message': f'cutover_scene 取值非法，可选值：{", ".join(sorted(WRITABLE_CUTOVER_SCENES))}'
            }), 400

        try:
            email_records_id = int(email_records_id)
        except (TypeError, ValueError):
            return jsonify({
                'success': False,
                'message': 'email_records_id 必须是整数'
            }), 400

        email_record = email_db.get_email_record_by_id(email_records_id)
        if not email_record:
            return jsonify({
                'success': False,
                'message': f'邮件记录不存在：email_records_id={email_records_id}'
            }), 404

        remark_text = normalize_excel_value(scene_remark) or ''
        if not email_db.update_email_record_cutover_scene(
            email_records_id, cutover_scene, remark_text,
        ):
            logger.warning(
                f"割接场景回写失败：email_records_id={email_records_id}, scene={cutover_scene}"
            )
            return jsonify({
                'success': False,
                'message': '割接场景写入失败'
            }), 500

        return jsonify({
            'success': True,
            'message': '割接场景回写成功',
            'data': {
                'email_records_id': email_records_id,
                'email_id': email_record.get('email_id'),
                'cutover_scene': cutover_scene,
                'cutover_scene_label': cutover_scene_label(cutover_scene),
                'scene_remark': remark_text,
            }
        }), 200
    except Exception as e:
        logger.error(f"割接场景回写失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/email/attachments/<path:relative_path>', methods=['GET'])
def download_email_attachment(relative_path):
    try:
        attachment_path = attachment_path_from_relative(relative_path)
        if not attachment_path or not attachment_path.exists() or not attachment_path.is_file():
            return jsonify({
                'success': False,
                'message': '附件不存在'
            }), 404

        return send_file(attachment_path, as_attachment=True)
    except Exception as e:
        logger.error(f"下载邮件附件失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


def render_admin_page():
    """提供管理台页面，并按 js/css 文件修改时间注入版本号避免浏览器缓存旧资源"""
    html = (BASE_DIR / 'static' / 'admin.html').read_text(encoding='utf-8')
    for name in ('admin.css', 'admin-shared.js', 'admin-i18n.js', 'admin-views.js', 'admin.js'):
        path = BASE_DIR / 'static' / name
        version = path.stat().st_mtime_ns
        html = html.replace(f'/static/{name}', f'/static/{name}?v={version}')
    return html


@app.route('/', methods=['GET'])
def dashboard_home():
    return render_admin_page()


@app.route('/admin', methods=['GET'])
def dashboard_admin():
    return render_admin_page()


@app.route('/admin/<path:subpath>', methods=['GET'])
def dashboard_admin_subpath(subpath):
    return render_admin_page()


@app.route('/api/email-records', methods=['GET'])
@require_api_key
def list_email_records():
    try:
        page = build_query_page(request.args.get('page'), request.args.get('pageSize'))
        records = get_record_query_repository().list_email_records(
            page,
            request.args.get('sender') or None,
            receiver=request.args.get('receiver') or None,
        )
        return jsonify({
            'success': True,
            'message': '邮件记录查询成功',
            'data': records
        }), 200
    except Exception as e:
        logger.error(f"查询邮件记录失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/auth/check', methods=['GET'])
@require_api_key
def check_admin_auth():
    """管理台登录校验：前端以 API Key 作为登录密码，校验通过即进入登录态。"""
    return jsonify({
        'success': True,
        'message': '认证通过',
    }), 200


@app.route('/api/dashboard/summary', methods=['GET'])
@require_api_key
def get_dashboard_summary():
    try:
        return jsonify({
            'success': True,
            'message': '运营摘要查询成功',
            'data': get_record_query_repository().get_operations_summary(),
        }), 200
    except Exception as e:
        logger.error(f"查询运营摘要失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/suppliers', methods=['GET'])
@require_api_key
def list_supplier_configs():
    try:
        records = get_supplier_config_repository().list()
        return jsonify({
            'success': True,
            'message': '供应商配置查询成功',
            'data': records
        }), 200
    except Exception as e:
        logger.error(f"查询供应商配置失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/suppliers', methods=['POST'])
@require_api_key
def create_supplier_config():
    try:
        payload = SupplierConfigCreate.model_validate(request.get_json(silent=True) or {})
        record = get_supplier_config_repository().create(payload)
        return jsonify({
            'success': True,
            'message': '供应商配置创建成功',
            'data': record
        }), 201
    except ValidationError as e:
        return jsonify({
            'success': False,
            'message': validation_error_message(e)
        }), 400
    except SupplierConfigConflictError as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 409
    except Exception as e:
        logger.error(f"创建供应商配置失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/suppliers/mail-types', methods=['GET'])
@require_api_key
def list_supplier_mail_types():
    """返回内置邮件类型列表，供前端配置邮件样本。"""
    return jsonify({
        'success': True,
        'message': '邮件类型查询成功',
        'data': list(MAIL_TYPE_OPTIONS)
    }), 200


@app.route('/api/suppliers/field-defaults', methods=['GET'])
@require_api_key
def list_supplier_field_defaults():
    """返回固定字段的内置默认提取规则，供前端预填。"""
    return jsonify({
        'success': True,
        'message': '固定字段默认规则查询成功',
        'data': get_fixed_field_definitions()
    }), 200


@app.route('/api/suppliers/preview-prompt', methods=['POST'])
@require_api_key
def preview_supplier_prompt():
    """根据提取字段配置预览自动生成的割接提示词。"""
    try:
        body = request.get_json(silent=True) or {}
        prompt = build_cutover_extract_prompt(
            body.get('line_custom_fields') or [],
            body.get('line_query_keywords') or [],
            body.get('extra_instructions') or '',
            body.get('fixed_field_rules') or {},
            custom_fields=body.get('custom_fields') or [],
        )
        return jsonify({
            'success': True,
            'message': '提示词预览生成成功',
            'data': {'cutover_extract_prompt': prompt}
        }), 200
    except ValidationError as e:
        return jsonify({
            'success': False,
            'message': validation_error_message(e)
        }), 400
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 400
    except Exception as e:
        logger.error(f"预览供应商提示词失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/suppliers/preview-attachments', methods=['POST'])
@require_api_key
def upload_supplier_preview_attachment():
    """上传提取预览用的附件，存储方式与监听邮件附件一致。"""
    try:
        uploaded_file = request.files.get('file')
        if uploaded_file is None or not uploaded_file.filename:
            return jsonify({
                'success': False,
                'message': '未接收到附件文件'
            }), 400
        relative_path = save_preview_attachment(
            uploaded_file.filename, uploaded_file.read()
        )
        return jsonify({
            'success': True,
            'message': '附件上传成功',
            'data': {
                'relative_path': relative_path,
                'url': build_attachment_url(relative_path),
            }
        }), 201
    except Exception as e:
        logger.error(f"上传提取预览附件失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/suppliers/preview-extract', methods=['POST'])
@require_api_key
def preview_supplier_extract():
    """调用 FastGPT 预览割接字段提取效果。

    请求格式与邮件转发 API（APIForwardAction）保持一致，
    转发的 content 中添加 extract_preview: true 表示提取预览任务。
    """
    try:
        body = request.get_json(silent=True) or {}
        attachments = []
        for relative_path in body.get('attachments') or []:
            if attachment_path_from_relative(relative_path) is None:
                return jsonify({
                    'success': False,
                    'message': f'附件不存在或路径非法: {relative_path}'
                }), 400
            attachments.append(relative_path)
        content_dict = {
            'extract_preview': True,
            'email_records_id': None,
            'email_id': None,
            'subject': body.get('subject') or '',
            'sender': body.get('sender') or '',
            'content': body.get('content') or '',
            'attachments': attachments,
            'attachment_urls': [
                build_attachment_url(relative_path)
                for relative_path in attachments
            ],
            'supplier_name': body.get('supplier_name') or None,
            'supplier_can_reply_directly': None,
            'supplier_cutover_extract_prompt': body.get('cutover_extract_prompt') or '',
            'supplier_mail_classify_prompt': build_mail_classify_prompt(
                body.get('supplier_name') or '',
                body.get('email_type_samples') or {},
            ),
            'securityTime': build_security_time(email_db),
        }
        if body.get('content_in_attachment'):
            content_dict['content_in_attachment'] = True

        api_request = APIRequest(
            chatId=str(uuid.uuid4()),
            stream=False,
            detail=False,
            messages=[{
                'content': json.dumps(content_dict, ensure_ascii=False, indent=2),
                'role': 'user',
            }],
        )
        headers = {
            'Authorization': f'Bearer {settings.api_token}',
            'Content-Type': 'application/json',
        }
        response = requests.post(
            settings.api_url,
            json=api_request.model_dump(),
            headers=headers,
            timeout=settings.api_timeout,
            verify=False,
        )
        if response.status_code not in (200, 202):
            logger.error(f"提取预览失败: {response.status_code} - {response.text}")
            return jsonify({
                'success': False,
                'message': f'FastGPT 调用失败: {response.status_code}'
            }), 502
        return jsonify({
            'success': True,
            'message': '提取预览成功',
            'data': {'extract_result': extract_fastgpt_reply(response.json())}
        }), 200
    except Exception as e:
        logger.error(f"提取预览异常: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/suppliers/<int:supplier_id>', methods=['GET'])
@require_api_key
def get_supplier_config(supplier_id):
    try:
        record = get_supplier_config_repository().get(supplier_id)
        if record is None:
            return jsonify({
                'success': False,
                'message': f'供应商配置不存在：id={supplier_id}'
            }), 404
        return jsonify({
            'success': True,
            'message': '供应商配置查询成功',
            'data': record
        }), 200
    except Exception as e:
        logger.error(f"查询供应商配置失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/suppliers/<int:supplier_id>', methods=['PATCH'])
@require_api_key
def update_supplier_config(supplier_id):
    try:
        payload = SupplierConfigUpdate.model_validate(request.get_json(silent=True) or {})
        record = get_supplier_config_repository().update(supplier_id, payload)
        if record is None:
            return jsonify({
                'success': False,
                'message': f'供应商配置不存在：id={supplier_id}'
            }), 404
        return jsonify({
            'success': True,
            'message': '供应商配置更新成功',
            'data': record
        }), 200
    except ValidationError as e:
        return jsonify({
            'success': False,
            'message': validation_error_message(e)
        }), 400
    except SupplierConfigConflictError as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 409
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 400
    except Exception as e:
        logger.error(f"更新供应商配置失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/suppliers/<int:supplier_id>', methods=['DELETE'])
@require_api_key
def delete_supplier_config(supplier_id):
    try:
        deleted = get_supplier_config_repository().delete(supplier_id)
        if not deleted:
            return jsonify({
                'success': False,
                'message': f'供应商配置不存在：id={supplier_id}'
            }), 404
        return jsonify({
            'success': True,
            'message': '供应商配置删除成功'
        }), 200
    except Exception as e:
        logger.error(f"删除供应商配置失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/tickets', methods=['GET'])
@require_api_key
def list_ticket_records():
    try:
        page = build_query_page(request.args.get('page'), request.args.get('pageSize'))
        records = get_record_query_repository().list_ticket_records(
            page,
            request.args.get('status') or None,
        )
        return jsonify({
            'success': True,
            'message': '工单记录查询成功',
            'data': records
        }), 200
    except Exception as e:
        logger.error(f"查询工单记录失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/tickets', methods=['POST'])
@require_api_key
def create_ticket():
    """
    新增工单记录。

    POST JSON:
        {
            "email_records_id": 1,
            "carrier_ticket_no": "RT123456",
            "cut_start_time": "2026-06-11 10:00:00",
            "cut_end_time": "2026-06-11 12:00:00",
            "status": "created",
            "cut_task_id": "CUT-001"
        }
    """
    try:
        data = request.get_json(silent=True) or {}

        email_records_id = data.get('email_records_id')
        carrier_ticket_no = data.get('carrier_ticket_no')
        cut_task_id = data.get('cut_task_id')
        status = data.get('status', 'created')

        if email_records_id is None or not carrier_ticket_no:
            return jsonify({
                'success': False,
                'message': '缺少必需参数：email_records_id, carrier_ticket_no'
            }), 400

        try:
            email_records_id = int(email_records_id)
        except (TypeError, ValueError):
            return jsonify({
                'success': False,
                'message': 'email_records_id 必须是整数'
            }), 400

        cut_start_time, error = parse_required_time(data.get('cut_start_time'), 'cut_start_time')
        if error:
            return jsonify({'success': False, 'message': error}), 400

        cut_end_time, error = parse_required_time(data.get('cut_end_time'), 'cut_end_time')
        if error:
            return jsonify({'success': False, 'message': error}), 400

        email_record = email_db.get_email_record_by_id(email_records_id)
        if not email_record:
            return jsonify({
                'success': False,
                'message': f'邮件记录不存在：email_records_id={email_records_id}'
            }), 404

        duplicate_ticket_record = email_db.find_duplicate_ticket_record(
            carrier_ticket_no=carrier_ticket_no,
            cut_start_time=cut_start_time,
            cut_end_time=cut_end_time,
        )

        ticket_id = email_db.add_ticket_record(
            email_records_id=email_records_id,
            carrier_ticket_no=carrier_ticket_no,
            cut_start_time=cut_start_time,
            cut_end_time=cut_end_time,
            status=status,
            cut_task_id=cut_task_id,
        )

        if not ticket_id:
            return jsonify({
                'success': False,
                'message': '新增工单记录失败'
            }), 500

        ticket_record = email_db.get_ticket_record(ticket_id)
        ticket_record['is_duplicate_ticket'] = duplicate_ticket_record is not None
        ticket_record['duplicate_ticket_record'] = duplicate_ticket_record

        return jsonify({
            'success': True,
            'message': '工单记录创建成功',
            'data': ticket_record
        }), 201

    except Exception as e:
        logger.error(f"新增工单记录失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/circuits/query', methods=['POST'])
@require_api_key
def query_circuits():
    """
    查询线路表。

    POST JSON:
        {
            "supplier": "RT",
            "keywords": ["751630", "1285332"]
        }
    """
    try:
        data = request.get_json(silent=True) or {}
        supplier = data.get('supplier')
        keywords = data.get('keywords')

        if not normalize_excel_value(supplier):
            return jsonify({
                'success': False,
                'message': '缺少必需参数：supplier'
            }), 400

        if not isinstance(keywords, list):
            return jsonify({
                'success': False,
                'message': 'keywords 必须是数组'
            }), 400

        results = query_supplier_circuits(supplier, keywords)

        return jsonify({
            'success': True,
            'message': '查询成功',
            'data': results
        }), 200
    except FileNotFoundError as e:
        logger.error(f"查询线路表失败: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    except Exception as e:
        logger.error(f"查询线路表失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


CIRCUIT_MANAGE_FIELDS = ('supplier', 'supplier_circuit_id', 'circuit_id', 'line_type', 'line_status', 'remark')


def _clean_circuit_text(value):
    """线路字段归一化：去首尾空白、统一换行，保留 / ; - 等名称内的特殊字符。"""
    if value is None:
        return ''
    return str(value).replace('\r\n', '\n').replace('\r', '\n').strip()


def _normalize_circuit_payload(payload):
    return {field: _clean_circuit_text(payload.get(field)) for field in CIRCUIT_MANAGE_FIELDS}


@app.route('/api/circuits', methods=['GET'])
@require_api_key
def list_circuits_admin():
    """分页查询线路表，支持供应商/类型/线路状态筛选与关键字模糊搜索。"""
    try:
        supplier = normalize_excel_value(request.args.get('supplier'))
        line_type = normalize_excel_value(request.args.get('line_type'))
        line_status = normalize_excel_value(request.args.get('line_status'))
        keyword = normalize_excel_value(request.args.get('keyword'))
        try:
            page = max(int(request.args.get('page', 1)), 1)
        except (TypeError, ValueError):
            page = 1
        try:
            page_size = min(max(int(request.args.get('page_size', 20)), 1), 200)
        except (TypeError, ValueError):
            page_size = 20

        result = email_db.list_supplier_circuits(
            supplier=supplier or None,
            line_type=line_type or None,
            line_status=line_status or None,
            keyword=keyword or None,
            limit=page_size,
            offset=(page - 1) * page_size,
        )

        return jsonify({
            'success': True,
            'message': '查询成功',
            'data': {
                'rows': result['rows'],
                'total': result['total'],
                'page': page,
                'page_size': page_size,
                'options': email_db.supplier_circuit_options(),
            }
        }), 200
    except Exception as e:
        logger.error(f"查询线路列表失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/circuits', methods=['POST'])
@require_api_key
def create_circuit_admin():
    """新增一条线路。"""
    try:
        fields = _normalize_circuit_payload(request.get_json(silent=True) or {})
        if not any(fields.values()):
            return jsonify({
                'success': False,
                'message': '所有字段均为空，无法新增线路'
            }), 400

        circuit_pk = email_db.create_supplier_circuit(fields)
        if circuit_pk is None:
            return jsonify({
                'success': False,
                'message': '新增线路失败'
            }), 500

        return jsonify({
            'success': True,
            'message': '新增线路成功',
            'data': email_db.get_supplier_circuit(circuit_pk)
        }), 201
    except Exception as e:
        logger.error(f"新增线路失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/circuits/<int:circuit_pk>', methods=['PATCH'])
@require_api_key
def update_circuit_admin(circuit_pk):
    """更新一条线路。"""
    try:
        payload = request.get_json(silent=True) or {}
        fields = {
            field: _clean_circuit_text(payload.get(field))
            for field in CIRCUIT_MANAGE_FIELDS
            if field in payload
        }
        if not fields:
            return jsonify({
                'success': False,
                'message': '未提供任何可更新字段'
            }), 400

        if not email_db.get_supplier_circuit(circuit_pk):
            return jsonify({
                'success': False,
                'message': f'线路不存在: id={circuit_pk}'
            }), 404

        if not email_db.update_supplier_circuit(circuit_pk, fields):
            return jsonify({
                'success': False,
                'message': '更新线路失败'
            }), 500

        return jsonify({
            'success': True,
            'message': '更新线路成功',
            'data': email_db.get_supplier_circuit(circuit_pk)
        }), 200
    except Exception as e:
        logger.error(f"更新线路失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/circuits/<int:circuit_pk>', methods=['DELETE'])
@require_api_key
def delete_circuit_admin(circuit_pk):
    """删除一条线路。"""
    try:
        if not email_db.delete_supplier_circuit(circuit_pk):
            return jsonify({
                'success': False,
                'message': f'线路不存在: id={circuit_pk}'
            }), 404

        return jsonify({
            'success': True,
            'message': '删除线路成功'
        }), 200
    except Exception as e:
        logger.error(f"删除线路失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/circuits/export', methods=['GET'])
def export_circuits_xlsx():
    """导出全量线路表为 xlsx 文件。"""
    try:
        rows = email_db.get_all_supplier_circuits()
        content = build_circuits_workbook_bytes(rows)
        return send_file(
            BytesIO(content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='线路表.xlsx'
        )
    except Exception as e:
        logger.error(f"导出线路表失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/circuits/import', methods=['POST'])
@require_api_key
def import_circuits_xlsx():
    """
    导入线路表 xlsx，全量替换模式。

    默认仅解析并返回预览（confirm != 'true'），前端展示预览并提示后，
    再次携带 confirm=true 提交才真正全量替换数据库中的线路数据。
    """
    try:
        uploaded = request.files.get('file')
        if uploaded is None or not uploaded.filename:
            return jsonify({
                'success': False,
                'message': '请上传 xlsx 文件（字段名 file）'
            }), 400

        try:
            rows, warnings = parse_circuits_workbook(BytesIO(uploaded.read()))
        except ValueError as e:
            return jsonify({
                'success': False,
                'message': str(e)
            }), 400
        except RuntimeError as e:
            return jsonify({
                'success': False,
                'message': str(e)
            }), 500

        confirm = str(request.form.get('confirm', '')).lower() == 'true'
        if not confirm:
            return jsonify({
                'success': True,
                'message': '解析成功，请确认预览后全量替换导入',
                'data': {
                    'total': len(rows),
                    'rows': rows[:50],
                    'warnings': warnings,
                }
            }), 200

        email_db.replace_all_supplier_circuits(rows)
        return jsonify({
            'success': True,
            'message': f'导入成功，已全量替换线路表，共 {len(rows)} 条',
            'data': {
                'total': len(rows),
                'warnings': warnings,
            }
        }), 200
    except Exception as e:
        logger.error(f"导入线路表失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500



@app.route('/api/template-xlsx', methods=['GET', 'POST'])
@require_api_key
def generate_template_xlsx():
    """
    生成 template.xlsx 中的两个业务 sheet：电路表、割接原因表。

    GET:
        生成模板中的两个 sheet，保存到 data 目录，并返回下载链接。

    POST JSON:
        {
            "filename": "template.xlsx",
            "circuits": [
                {
                    "客户名称": "...",
                    "电路代号": "..."
                }
            ],
            "reasons": [
                {
                    "割接线路/设备名称": "...",
                    "割接原因": "..."
                }
            ]
        }
    """
    try:
        payload = request.get_json(silent=True) if request.method == 'POST' else None
        output = build_template_workbook(payload)
        filename = 'template.xlsx'

        if payload and payload.get('filename'):
            filename = payload['filename']

        output_path = save_workbook_output(output, filename)
        download_url = url_for(
            'download_template_xlsx',
            filename=output_path.name,
            _external=True
        )

        logger.info(f"生成 Excel 文件成功: {output_path}")
        return jsonify({
            'success': True,
            'message': 'Excel 文件生成成功',
            'data': {
                'filename': output_path.name,
                'file_path': str(output_path),
                'download_url': download_url
            }
        }), 200
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 400
    except FileNotFoundError as e:
        logger.error(f"生成 Excel 文件失败: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    except Exception as e:
        logger.error(f"生成 Excel 文件失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/tasks/generate', methods=['POST'])
@app.route('/api/cutover/fill', methods=['POST'])  # 旧路径兼容，FastGPT 存量工作流仍可使用
@require_api_key
def generate_cutover_tasks():
    """
    根据割接入参生成填报结果并创建割接上报任务。

    客户线路：复用 template.xlsx 的电路表、割接原因表生成 Excel。
    骨干线路：仅返回填报字段，不生成 Excel。
    必须传入 email_records_id 关联来源邮件，成功后自动创建/更新割接上报任务。
    """
    try:
        payload = request.get_json(silent=True) or {}

        email_records_id = payload.get('email_records_id')
        if email_records_id is None:
            return jsonify({
                'success': False,
                'message': '缺少必需参数：email_records_id'
            }), 400
        try:
            email_records_id = int(email_records_id)
        except (TypeError, ValueError):
            return jsonify({
                'success': False,
                'message': 'email_records_id 必须是整数'
            }), 400
        if not email_db.get_email_record_by_id(email_records_id):
            return jsonify({
                'success': False,
                'message': f'邮件记录不存在：email_records_id={email_records_id}'
            }), 404

        result = build_cutover_fill_response(payload)
        customer_excel = None

        if result.get('circuits'):
            output = build_template_workbook(result)
            output_path = save_workbook_output(output, result.get('filename'))
            customer_excel = {
                'filename': output_path.name,
                'download_url': url_for(
                    'download_template_xlsx',
                    filename=output_path.name,
                    _external=True
                )
            }

        tasks = save_cutover_fill_tasks(
            email_records_id=email_records_id,
            payload=payload,
            result=result,
            customer_excel_filename=customer_excel['filename'] if customer_excel else None,
        )

        hidden_response_keys = {'filename', 'circuits', 'reasons'}

        response_data = {
            key: value
            for key, value in result.items()
            if key not in hidden_response_keys
        }
        response_data['customer_excel'] = customer_excel
        response_data['tasks'] = [
            {
                'id': task.get('id'),
                'line_type': task.get('line_type'),
                'line_type_label': cutover_line_type_label(task.get('line_type')),
                'status': task.get('status'),
            }
            for task in tasks
        ]
        response_data['msg'] = build_cutover_fill_msg(response_data, email_records_id)

        return jsonify({
            'success': True,
            'message': '割接填报结果生成成功',
            'data': response_data
        }), 200
    except ValueError as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 400
    except FileNotFoundError as e:
        logger.error(f"生成割接填报结果失败: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    except Exception as e:
        logger.error(f"生成割接填报结果失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/template-xlsx/download/<path:filename>', methods=['GET'])
def download_template_xlsx(filename):
    try:
        safe_filename = normalize_xlsx_filename(filename)
        file_path = DATA_DIR / safe_filename

        if not file_path.exists() or not file_path.is_file():
            return jsonify({
                'success': False,
                'message': f'文件不存在: {safe_filename}'
            }), 404

        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=safe_filename
        )
    except Exception as e:
        logger.error(f"下载 Excel 文件失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/tasks', methods=['GET'])
@require_api_key
def list_cutover_tasks():
    """分页查询割接上报任务，支持按状态、供应商过滤。"""
    try:
        page = build_query_page(request.args.get('page'), request.args.get('pageSize'))
        records = get_record_query_repository().list_cutover_tasks(
            page,
            request.args.get('status') or None,
            request.args.get('supplier') or None,
        )
        for item in records.get('items', []):
            item['status_label'] = cutover_task_status_label(item.get('status'))
            item['line_type_label'] = cutover_line_type_label(item.get('line_type'))
        return jsonify({
            'success': True,
            'message': '割接任务查询成功',
            'data': records
        }), 200
    except Exception as e:
        logger.error(f"查询割接任务失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/emails/mail-types', methods=['GET'])
@require_api_key
def list_cutover_email_mail_types():
    """返回已解析的邮件类型去重列表，供割接列表标签筛选下拉使用。"""
    try:
        mail_types = get_record_query_repository().get_distinct_mail_types()
        return jsonify({
            'success': True,
            'message': '邮件类型查询成功',
            'data': mail_types,
        }), 200
    except Exception as e:
        logger.error(f"查询邮件类型失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/cutover/emails', methods=['GET'])
@require_api_key
def list_cutover_task_emails():
    """按邮件维度分页查询供应商邮件及其割接任务，每封邮件内嵌其割接任务列表。"""
    try:
        tag = request.args.get('tag') or None
        if tag is not None and tag not in VALID_CUTOVER_TAGS:
            return jsonify({
                'success': False,
                'message': f'tag 取值非法，可选值：{", ".join(sorted(VALID_CUTOVER_TAGS))}'
            }), 400
        page = build_query_page(request.args.get('page'), request.args.get('pageSize'))
        records = get_record_query_repository().list_cutover_task_emails(
            page,
            request.args.get('status') or None,
            request.args.get('supplier') or None,
            sender=request.args.get('sender') or None,
            start_time=request.args.get('start') or None,
            end_time=request.args.get('end') or None,
            tag=tag,
            mail_type=request.args.get('mail_type') or None,
            receiver=request.args.get('receiver') or None,
        )
        for item in records.get('items', []):
            for task in item.get('tasks', []):
                task['status_label'] = cutover_task_status_label(task.get('status'))
                task['line_type_label'] = cutover_line_type_label(task.get('line_type'))
        return jsonify({
            'success': True,
            'message': '割接任务邮件查询成功',
            'data': records
        }), 200
    except Exception as e:
        logger.error(f"查询割接任务邮件失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/emails/<int:email_records_id>', methods=['GET'])
@require_api_key
def get_cutover_email_detail(email_records_id):
    """获取割接邮件详情：邮件正文、附件及其下的割接任务列表。"""
    try:
        detail = get_record_query_repository().get_cutover_email_detail(email_records_id)
        if not detail:
            return jsonify({
                'success': False,
                'message': f'邮件记录不存在：id={email_records_id}'
            }), 404
        for task in detail.get('tasks', []):
            task['status_label'] = cutover_task_status_label(task.get('status'))
            task['line_type_label'] = cutover_line_type_label(task.get('line_type'))

        # 基于任务保存的填报入参实时重查线路表，展示当前匹配到的线路数据
        line_table = None
        for task in detail.get('tasks', []):
            full_task = email_db.get_cutover_task(task.get('id'))
            fill_payload = (full_task or {}).get('fill_payload')
            if not fill_payload:
                continue
            try:
                line_table = match_supplier_circuits_by_payload(fill_payload)
            except Exception as match_error:
                logger.warning(f"邮件详情重查线路表失败: email_records_id={email_records_id}, {match_error}")
                line_table = {'error': str(match_error)}
            break
        detail['line_table'] = line_table

        return jsonify({
            'success': True,
            'message': '割接邮件详情查询成功',
            'data': detail
        }), 200
    except Exception as e:
        logger.error(f"查询割接邮件详情失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/tasks/<int:task_id>', methods=['GET'])
@require_api_key
def get_cutover_task(task_id):
    """获取割接任务详情，含填报数据、关联邮件与上报记录。"""
    try:
        task = get_cutover_task_detail(task_id)
        if not task:
            return jsonify({
                'success': False,
                'message': f'割接任务不存在：id={task_id}'
            }), 404

        if task.get('customer_excel_filename'):
            task['excel_download_url'] = url_for(
                'download_cutover_task_excel',
                task_id=task_id,
                _external=True
            )

        return jsonify({
            'success': True,
            'message': '割接任务详情查询成功',
            'data': task
        }), 200
    except Exception as e:
        logger.error(f"查询割接任务详情失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/tasks/<int:task_id>', methods=['PATCH'])
@require_api_key
def edit_cutover_task(task_id):
    """
    编辑割接任务的填报数据。

    PATCH JSON（均可选，传哪个改哪个）:
        {
            "circuits": [...],          # 客户线路电路表行
            "reasons": [...],           # 客户线路割接原因表行
            "backbone_circuits": [...]  # 骨干线路填报字段
        }

    编辑后任务状态回退为 draft，客户侧数据变更会自动重新生成 Excel。
    """
    try:
        data = request.get_json(silent=True) or {}
        edits = {
            key: data[key]
            for key in ('circuits', 'reasons', 'backbone_circuits')
            if key in data
        }
        if not edits:
            return jsonify({
                'success': False,
                'message': '未提供可编辑字段：circuits / reasons / backbone_circuits'
            }), 400

        task = apply_task_edit(task_id, edits)
        return jsonify({
            'success': True,
            'message': '割接任务已更新，状态回退为待确认',
            'data': task
        }), 200
    except LookupError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except FileNotFoundError as e:
        logger.error(f"编辑割接任务重新生成 Excel 失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    except Exception as e:
        logger.error(f"编辑割接任务失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/tasks/<int:task_id>/switch-type', methods=['POST'])
@require_api_key
def switch_cutover_task_type(task_id):
    """
    切换割接任务的线路类型（客户线路 <-> 骨干线路）。

    已有填报数据会尽量转换带入新类型的填报；切换后任务状态回退为 draft，
    切换为客户线路时自动重新生成 Excel。同一封邮件已存在目标类型的任务时拒绝切换。

    POST JSON:
        {"line_type": "customer" | "backbone"}
    """
    try:
        data = request.get_json(silent=True) or {}
        task = switch_task_line_type(task_id, data.get('line_type'))
        task['status_label'] = cutover_task_status_label(task.get('status'))
        task['line_type_label'] = cutover_line_type_label(task.get('line_type'))
        return jsonify({
            'success': True,
            'message': f"任务已切换为{cutover_line_type_label(task.get('line_type'))}填报",
            'data': task
        }), 200
    except LookupError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except FileNotFoundError as e:
        logger.error(f"切换割接任务类型失败: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    except Exception as e:
        logger.error(f"切换割接任务类型失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/tasks/<int:task_id>/confirm', methods=['POST'])
@require_api_key
def confirm_cutover_task(task_id):
    """人工确认割接任务：draft -> confirmed。"""
    try:
        task = confirm_task(task_id)
        task['status_label'] = cutover_task_status_label(task.get('status'))
        return jsonify({
            'success': True,
            'message': '割接任务已确认，可以发起上报',
            'data': task
        }), 200
    except LookupError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        logger.error(f"确认割接任务失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/tasks/<int:task_id>/report', methods=['POST'])
@require_api_key
def report_cutover_task_route(task_id):
    """
    触发割接任务上报综调系统。

    当前 Playwright 对接尚未实现：会写入一条 pending 状态的上报记录，
    任务保持在已确认状态，可重复触发。对接完成后替换 cutover_report.py
    中的 report_cutover_task 实现即可。

    POST JSON（可选）:
        {"report_type": "customer" | "backbone"}

    不传 report_type 时默认使用任务的线路类型。
    """
    try:
        task = email_db.get_cutover_task(task_id)
        if not task:
            return jsonify({
                'success': False,
                'message': f'割接任务不存在：id={task_id}'
            }), 404

        if task.get('status') not in ('confirmed', 'report_failed'):
            return jsonify({
                'success': False,
                'message': f"仅已确认或上报失败的任务可以上报，当前状态：{cutover_task_status_label(task.get('status'))}"
            }), 400

        data = request.get_json(silent=True) or {}
        report_type = data.get('report_type') or task.get('line_type')
        report_id = email_db.add_cutover_report(
            task_id,
            report_type=report_type,
            status='pending',
            result={'note': '等待上报执行'},
        )
        if not report_id:
            return jsonify({
                'success': False,
                'message': '新增上报记录失败'
            }), 500

        try:
            outcome = report_cutover_task(task)
        except CutoverReportNotReady as e:
            email_db.update_cutover_report(report_id, result={'note': str(e)})
            return jsonify({
                'success': True,
                'message': str(e),
                'data': {
                    'report_id': report_id,
                    'report_status': 'pending',
                    'task_status': task.get('status'),
                }
            }), 202

        reported = (outcome or {}).get('status') == 'success'
        email_db.update_cutover_report(report_id, status='success' if reported else 'failed', result=outcome)
        email_db.update_cutover_task(
            task_id,
            status='reported' if reported else 'report_failed',
        )
        return jsonify({
            'success': True,
            'message': '上报成功' if reported else '上报失败',
            'data': {
                'report_id': report_id,
                'report_status': 'success' if reported else 'failed',
                'task_status': 'reported' if reported else 'report_failed',
                'result': outcome,
            }
        }), 200
    except Exception as e:
        logger.error(f"上报割接任务失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/cutover/tasks/<int:task_id>/excel', methods=['GET'])
def download_cutover_task_excel(task_id):
    """下载割接任务当前版本的客户线路 Excel。"""
    try:
        task = email_db.get_cutover_task(task_id)
        if not task:
            return jsonify({
                'success': False,
                'message': f'割接任务不存在：id={task_id}'
            }), 404

        filename = task.get('customer_excel_filename')
        if not filename:
            return jsonify({
                'success': False,
                'message': f'该任务未生成客户线路 Excel：id={task_id}'
            }), 404

        safe_filename = normalize_xlsx_filename(filename)
        file_path = DATA_DIR / safe_filename
        if not file_path.exists() or not file_path.is_file():
            return jsonify({
                'success': False,
                'message': f'文件不存在: {safe_filename}'
            }), 404

        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=safe_filename
        )
    except Exception as e:
        logger.error(f"下载割接任务 Excel 失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500



# ---------- 系统配置 ----------

GUARD_TIME_FORMAT = '%Y-%m-%d %H:%M:%S'

GUARD_TIME_KEYS = ('guard_start_time', 'guard_end_time')


def is_valid_guard_time(value: str) -> bool:
    try:
        datetime.strptime(value, GUARD_TIME_FORMAT)
        return True
    except ValueError:
        return False


def get_guard_time_settings() -> dict:
    stored = email_db.get_system_settings()
    return {key: stored.get(key) for key in GUARD_TIME_KEYS}


@app.route('/api/system/settings', methods=['GET'])
@require_api_key
def get_system_settings():
    try:
        return jsonify({
            'success': True,
            'message': '系统配置查询成功',
            'data': get_guard_time_settings()
        }), 200
    except Exception as e:
        logger.error(f"查询系统配置失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


@app.route('/api/system/settings', methods=['PUT'])
@require_api_key
def update_system_settings():
    try:
        payload = request.get_json(silent=True) or {}
        updates = {}
        for key in GUARD_TIME_KEYS:
            if key not in payload:
                continue
            value = payload.get(key)
            if value in (None, ''):
                updates[key] = None
                continue
            if not isinstance(value, str) or not is_valid_guard_time(value.strip()):
                return jsonify({
                    'success': False,
                    'message': f'时间格式错误（{key}），应为 YYYY-MM-DD HH:MM:SS，如 2026-07-15 00:00:00'
                }), 400
            updates[key] = value.strip()

        if not updates:
            return jsonify({
                'success': False,
                'message': '未提供可更新的配置项'
            }), 400

        for key, value in updates.items():
            email_db.set_system_setting(key, value)

        return jsonify({
            'success': True,
            'message': '系统配置保存成功',
            'data': get_guard_time_settings()
        }), 200
    except Exception as e:
        logger.error(f"更新系统配置失败: {e}")
        return jsonify({
            'success': False,
            'message': f'服务器错误: {str(e)}'
        }), 500


# ---------- 邮箱账号配置（多邮箱页面化管理） ----------

def build_reply_email_client(email_record: dict):
    """根据邮件记录的收件账号匹配邮箱配置；找不到时回退第一个启用账号（兼容存量记录）。"""
    receiver = (email_record.get('receiver') or '').strip()
    if receiver:
        row = email_db.get_mail_account_by_address(receiver)
        if row and row.get('enabled', 1):
            config = account_row_to_config(row)
            if config:
                return EmailClient(config)
    configs = list_enabled_account_configs()
    if configs:
        return EmailClient(configs[0])
    return None


def _mail_account_to_api(row: dict) -> dict:
    """数据库行转接口返回结构（密码脱敏为 password_set 标记）。"""
    return {
        'id': row['id'],
        'name': row.get('name') or '',
        'email_address': row['email_address'],
        'password_set': bool(row.get('password_enc')),
        'imap_server': row['imap_server'],
        'imap_port': row.get('imap_port') or 993,
        'imap_use_ssl': bool(row.get('imap_use_ssl', 1)),
        'smtp_server': row.get('smtp_server') or '',
        'smtp_port': row.get('smtp_port') or 465,
        'smtp_use_ssl': bool(row.get('smtp_use_ssl', 1)),
        'smtp_use_tls': bool(row.get('smtp_use_tls', 0)),
        'enabled': bool(row.get('enabled', 1)),
        'create_time': row.get('create_time'),
        'update_time': row.get('update_time'),
    }


def _parse_mail_account_payload(payload: dict, require_password: bool):
    """解析并校验邮箱账号请求体，返回 (fields, error_message)。"""
    fields = {}
    error = None

    if 'name' in payload:
        fields['name'] = str(payload.get('name') or '').strip()

    if 'email_address' in payload:
        address = str(payload.get('email_address') or '').strip()
        if not address:
            error = 'email_address 不能为空'
        else:
            fields['email_address'] = address

    password = payload.get('email_password')
    if password is not None:
        password = str(password).strip()
        if password:
            fields['password_enc'] = encrypt_password(password)
        elif require_password:
            error = error or 'email_password 不能为空'

    if 'imap_server' in payload:
        imap_server = str(payload.get('imap_server') or '').strip()
        if not imap_server:
            error = error or 'imap_server 不能为空'
        else:
            fields['imap_server'] = imap_server

    for port_key, default in (('imap_port', 993), ('smtp_port', 465)):
        if port_key in payload and payload.get(port_key) not in (None, ''):
            try:
                fields[port_key] = int(payload.get(port_key))
            except (TypeError, ValueError):
                error = error or f'{port_key} 必须是整数'

    for bool_key in ('imap_use_ssl', 'smtp_use_ssl', 'smtp_use_tls', 'enabled'):
        if bool_key in payload and payload.get(bool_key) is not None:
            fields[bool_key] = 1 if payload.get(bool_key) else 0

    if 'smtp_server' in payload:
        fields['smtp_server'] = str(payload.get('smtp_server') or '').strip()

    if require_password:
        for key in ('name', 'email_address', 'imap_server'):
            if not fields.get(key):
                error = error or f'{key} 不能为空'

    return fields, error


@app.route('/api/system/mail-accounts', methods=['GET'])
@require_api_key
def list_mail_accounts_api():
    """查询全部邮箱账号（密码脱敏）。"""
    try:
        rows = email_db.list_mail_accounts()
        return jsonify({
            'success': True,
            'message': '邮箱账号查询成功',
            'data': [_mail_account_to_api(row) for row in rows]
        }), 200
    except Exception as e:
        logger.error(f"查询邮箱账号失败: {e}")
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'}), 500


def _sync_listeners_with_warning(base_message: str) -> str:
    """热同步监听；部分邮箱启动失败时在提示语中追加警告（账号已保存不受影响）。"""
    failed = listener_manager.sync()
    if isinstance(failed, list) and failed:
        return f"{base_message}，但监听启动失败：{'、'.join(failed)}，请检查授权码与服务器设置"
    return base_message


@app.route('/api/system/mail-accounts', methods=['POST'])
@require_api_key
def create_mail_account_api():
    """新建邮箱账号，成功后热同步监听。"""
    try:
        payload = request.get_json(silent=True) or {}
        fields, error = _parse_mail_account_payload(payload, require_password=True)
        if error:
            return jsonify({'success': False, 'message': error}), 400

        account_id = email_db.add_mail_account(fields)
        if account_id is None:
            return jsonify({
                'success': False,
                'message': f"新建失败：邮箱地址 {fields.get('email_address')} 可能已存在"
            }), 400

        message = _sync_listeners_with_warning('邮箱账号新建成功')
        row = email_db.get_mail_account(account_id)
        return jsonify({
            'success': True,
            'message': message,
            'data': _mail_account_to_api(row)
        }), 201
    except Exception as e:
        logger.error(f"新建邮箱账号失败: {e}")
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'}), 500


@app.route('/api/system/mail-accounts/<int:account_id>', methods=['PUT'])
@require_api_key
def update_mail_account_api(account_id):
    """更新邮箱账号（密码传空串表示保持不变），成功后热同步监听。"""
    try:
        existing = email_db.get_mail_account(account_id)
        if not existing:
            return jsonify({'success': False, 'message': '邮箱账号不存在'}), 404

        payload = request.get_json(silent=True) or {}
        fields, error = _parse_mail_account_payload(payload, require_password=False)
        if error:
            return jsonify({'success': False, 'message': error}), 400
        if not fields:
            return jsonify({'success': False, 'message': '未提供可更新的字段'}), 400

        if not email_db.update_mail_account(account_id, fields):
            return jsonify({
                'success': False,
                'message': f"更新失败：邮箱地址 {fields.get('email_address')} 可能与其他账号冲突"
            }), 400

        message = _sync_listeners_with_warning('邮箱账号更新成功')
        row = email_db.get_mail_account(account_id)
        return jsonify({
            'success': True,
            'message': message,
            'data': _mail_account_to_api(row)
        }), 200
    except Exception as e:
        logger.error(f"更新邮箱账号失败: {e}")
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'}), 500


@app.route('/api/system/mail-accounts/<int:account_id>', methods=['DELETE'])
@require_api_key
def delete_mail_account_api(account_id):
    """删除邮箱账号，同步停止对应监听。"""
    try:
        if not email_db.delete_mail_account(account_id):
            return jsonify({'success': False, 'message': '邮箱账号不存在'}), 404
        listener_manager.sync()
        return jsonify({'success': True, 'message': '邮箱账号删除成功'}), 200
    except Exception as e:
        logger.error(f"删除邮箱账号失败: {e}")
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'}), 500


@app.route('/api/system/mail-accounts/status', methods=['GET'])
@require_api_key
def mail_accounts_status_api():
    """查询各邮箱账号监听状态：运行中/失败原因 + 最后收件时间（无实时收件时回退历史记录）。"""
    try:
        status = listener_manager.get_status()
        if not isinstance(status, dict):
            status = {}
        result = {}
        for row in email_db.list_mail_accounts():
            item = dict(status.get(row['id']) or {})
            item.setdefault('running', False)
            item.setdefault('error', None)
            if not item.get('last_email_at'):
                item['last_email_at'] = email_db.get_last_email_time_by_receiver(row['email_address'])
            result[row['id']] = item
        return jsonify({
            'success': True,
            'message': '监听状态查询成功',
            'data': result,
        }), 200
    except Exception as e:
        logger.error(f"查询邮箱监听状态失败: {e}")
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'}), 500


@app.route('/api/system/mail-accounts/test-preview', methods=['POST'])
@require_api_key
def test_mail_account_preview_api():
    """保存前用表单配置测试 IMAP 连接（不写入数据库）。"""
    try:
        payload = request.get_json(silent=True) or {}
        fields, error = _parse_mail_account_payload(payload, require_password=True)
        if error:
            return jsonify({'success': False, 'message': error}), 400

        config = MailAccountConfig(
            id=0,
            name=fields.get('name') or '',
            email_address=fields['email_address'],
            email_password=str(payload.get('email_password') or '').strip(),
            imap_server=fields['imap_server'],
            imap_port=fields.get('imap_port') or 993,
            imap_use_ssl=bool(fields.get('imap_use_ssl', 1)),
            smtp_server=fields.get('smtp_server') or '',
            smtp_port=fields.get('smtp_port') or 465,
            smtp_use_ssl=bool(fields.get('smtp_use_ssl', 1)),
            smtp_use_tls=bool(fields.get('smtp_use_tls', 0)),
        )
        client = EmailClient(config)
        if client.connect():
            client.disconnect()
            message = f'IMAP 连接成功: {config.imap_server}:{config.imap_port}'
            return jsonify({
                'success': True,
                'message': message,
                'data': {'message': message},
            }), 200
        return jsonify({
            'success': False,
            'message': f'IMAP 连接失败，请检查服务器/端口/账号/密码: {config.imap_server}:{config.imap_port}'
        }), 400
    except Exception as e:
        logger.error(f"测试邮箱连接（预览）失败: {e}")
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'}), 500


@app.route('/api/system/mail-accounts/<int:account_id>/test', methods=['POST'])
@require_api_key
def test_mail_account_api(account_id):
    """测试邮箱账号 IMAP 连接（登录并断开）。"""
    try:
        row = email_db.get_mail_account(account_id)
        if not row:
            return jsonify({'success': False, 'message': '邮箱账号不存在'}), 404

        config = account_row_to_config(row)
        if config is None:
            return jsonify({
                'success': False,
                'message': '密码缺失或解密失败，请先重新设置密码'
            }), 400

        client = EmailClient(config)
        ok = client.connect()
        if ok:
            client.disconnect()
            message = f'IMAP 连接成功: {config.imap_server}:{config.imap_port}'
            return jsonify({
                'success': True,
                'message': message,
                'data': {'message': message}
            }), 200
        return jsonify({
            'success': False,
            'message': f'IMAP 连接失败，请检查服务器/端口/账号/密码: {config.imap_server}:{config.imap_port}'
        }), 400
    except Exception as e:
        logger.error(f"测试邮箱连接失败: {e}")
        return jsonify({'success': False, 'message': f'服务器错误: {str(e)}'}), 500



if __name__ == '__main__':
    
    logger.info("启动 API 服务...")
    logger.info(f"API 地址: http://0.0.0.0:{settings.api_port}")
    logger.info("API Key 使用方式:")
    logger.info("  Header: Authorization: Bearer <your-api-key>")

    app.run(host='0.0.0.0', port=settings.api_port, debug=True)
