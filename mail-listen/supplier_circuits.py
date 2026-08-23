"""
线路表（supplier_circuits）业务模块：
- 解析导入的 Excel 线路表（按表头名匹配列，兼容旧版表头）
- 导出线路表 Excel
- 首次启动时将 data/线路表.xlsx 迁移进数据库
"""
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from loguru import logger

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError:
    Workbook = None
    load_workbook = None

from database import email_db

BASE_DIR = Path(__file__).resolve().parent
SUPPLIER_CIRCUITS_XLSX_PATH = BASE_DIR / 'data' / '线路表.xlsx'

# 导出表头顺序，与线路表 Excel 保持一致
CIRCUIT_EXPORT_HEADERS = ('Supplier', 'Supplier Circuit ID', 'Circuit ID', '类型', '线路状态', '备注')
CIRCUIT_EXPORT_SHEET_NAME = '线路表'

# 表头名（小写去空格）到字段的映射，兼容旧版「客户/骨干」表头
CIRCUIT_HEADER_ALIASES = {
    'supplier': 'supplier',
    'suppliercircuitid': 'supplier_circuit_id',
    'circuitid': 'circuit_id',
    '类型': 'line_type',
    '客户/骨干': 'line_type',
    '线路状态': 'line_status',
    '备注': 'remark',
}
CIRCUIT_REQUIRED_HEADERS = ('supplier', 'supplier_circuit_id', 'circuit_id')


def normalize_circuit_cell(value: Any) -> str:
    """单元格归一化：数字去小数尾巴、换行统一为 \\n、去除首尾空白。"""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    return text.strip()


def _map_header_row(cells: List[Any]) -> Dict[int, str]:
    """识别表头行，返回 列下标 -> 字段名 映射。"""
    mapping: Dict[int, str] = {}
    for index, cell in enumerate(cells):
        header = normalize_circuit_cell(cell).replace(' ', '')
        field = CIRCUIT_HEADER_ALIASES.get(header.lower()) or CIRCUIT_HEADER_ALIASES.get(header)
        if field and field not in mapping.values():
            mapping[index] = field
    return mapping


def parse_circuits_workbook(source) -> Tuple[List[Dict[str, str]], List[str]]:
    """
    解析线路表 Excel（文件路径或二进制流）。

    返回 (rows, warnings)：rows 为字段字典列表，warnings 为解析提示。
    表头无法识别或完全无数据行时抛出 ValueError。
    """
    if load_workbook is None:
        raise RuntimeError('缺少依赖 openpyxl，无法解析线路表')

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_cells = next(rows_iter, None)
        if header_cells is None:
            raise ValueError('线路表为空，未找到表头行')

        header_map = _map_header_row(list(header_cells))
        missing = [field for field in CIRCUIT_REQUIRED_HEADERS if field not in header_map.values()]
        if missing:
            raise ValueError(f'线路表表头缺少必需列：{", ".join(missing)}')

        rows: List[Dict[str, str]] = []
        warnings: List[str] = []
        empty_rows = 0

        for raw_row in rows_iter:
            if raw_row is None:
                continue
            row_data = {field: '' for field in CIRCUIT_HEADER_ALIASES.values()}
            for index, field in header_map.items():
                if index < len(raw_row):
                    row_data[field] = normalize_circuit_cell(raw_row[index])
            if not any(row_data.values()):
                empty_rows += 1
                continue
            rows.append(row_data)

        if empty_rows:
            warnings.append(f'跳过 {empty_rows} 行空行')
        if not rows:
            raise ValueError('线路表中没有有效数据行')
        return rows, warnings
    finally:
        wb.close()


def build_circuits_workbook_bytes(rows: List[Dict[str, Any]]) -> bytes:
    """把线路数据导出为 xlsx 字节流，Circuit ID 等字段保留换行原文。"""
    if Workbook is None:
        raise RuntimeError('缺少依赖 openpyxl，无法导出线路表')

    wb = Workbook()
    ws = wb.active
    ws.title = CIRCUIT_EXPORT_SHEET_NAME
    ws.append(list(CIRCUIT_EXPORT_HEADERS))

    field_order = ('supplier', 'supplier_circuit_id', 'circuit_id', 'line_type', 'line_status', 'remark')
    for row in rows:
        ws.append([row.get(field) or '' for field in field_order])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def migrate_supplier_circuits_from_xlsx(
    db=None,
    xlsx_path: Optional[Path] = None,
) -> int:
    """首次启动迁移：线路表数据库为空且存在 xlsx 文件时，把文件数据导入数据库。"""
    db = db or email_db
    xlsx_path = Path(xlsx_path) if xlsx_path else SUPPLIER_CIRCUITS_XLSX_PATH

    if db.count_supplier_circuits() > 0:
        return 0
    if not xlsx_path.exists():
        logger.warning(f"线路表数据库为空且文件不存在，跳过迁移: {xlsx_path}")
        return 0

    try:
        rows, warnings = parse_circuits_workbook(xlsx_path)
    except Exception as e:
        logger.error(f"从 {xlsx_path} 迁移线路表失败: {e}")
        return 0

    for warning in warnings:
        logger.warning(f"线路表迁移提示: {warning}")
    db.replace_all_supplier_circuits(rows)
    logger.info(f"线路表迁移完成: {xlsx_path} -> 数据库，共 {len(rows)} 条")
    return len(rows)
