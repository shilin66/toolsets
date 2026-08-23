import io
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


os.environ.setdefault('IMAP_SERVER', 'imap.example.com')
os.environ.setdefault('EMAIL_ADDRESS', 'test@example.com')
os.environ.setdefault('EMAIL_PASSWORD', 'password')
os.environ.setdefault('API_URL', 'https://api.example.com')
os.environ.setdefault('API_TOKEN', 'token')
os.environ.setdefault('API_KEY', 'test-key')

sys.path.insert(0, str(Path(__file__).resolve().parent))

import api_server
import supplier_circuits
from database import EmailDatabase


MULTI_LINE_CIRCUIT_ID = (
    'Belgorsk/RT-STO/CTEUR 16V4C001 Main(1st 2.5G TEA)\n'
    'Belgorsk/RT-STO/CTEUR 16V4C001 Spare(1st 2.5G TEA备)\n'
    'Belgorsk/RT-STO/CTEUR 16V4C002 Main(2nd 2.5G TEA)'
)


def build_import_workbook_bytes(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = '线路表'
    ws.append(['Supplier', 'Supplier Circuit ID', 'Circuit ID', '类型', '线路状态', '备注'])
    for row in rows:
        ws.append([
            row.get('supplier', ''),
            row.get('supplier_circuit_id', ''),
            row.get('circuit_id', ''),
            row.get('line_type', ''),
            row.get('line_status', ''),
            row.get('remark', ''),
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class SupplierCircuitsApiTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.db = EmailDatabase(os.path.join(self.tmpdir.name, 'mail_listener.db'))
        self.api_db_patch = patch.object(api_server, 'email_db', self.db)
        self.api_db_patch.start()
        self.client = api_server.app.test_client()
        self.headers = {'Authorization': 'Bearer test-key'}

    def tearDown(self):
        self.api_db_patch.stop()
        self.tmpdir.cleanup()

    def create_circuit(self, **overrides):
        payload = {
            'supplier': 'RT',
            'supplier_circuit_id': '1160515',
            'circuit_id': MULTI_LINE_CIRCUIT_ID,
            'line_type': '骨干',
            'line_status': '正常',
            'remark': '不中断割接直接上报四条中继线路；中断割接需按客户电路中断上报',
        }
        payload.update(overrides)
        response = self.client.post('/api/circuits', headers=self.headers, json=payload)
        self.assertEqual(response.status_code, 201, response.get_data(as_text=True))
        return response.get_json()['data']

    def test_create_and_get_list(self):
        created = self.create_circuit()
        self.assertEqual(created['supplier'], 'RT')
        self.assertEqual(created['circuit_id'], MULTI_LINE_CIRCUIT_ID)

        response = self.client.get('/api/circuits', headers=self.headers)
        self.assertEqual(response.status_code, 200)
        data = response.get_json()['data']
        self.assertEqual(data['total'], 1)
        self.assertEqual(data['rows'][0]['circuit_id'], MULTI_LINE_CIRCUIT_ID)
        self.assertIn('RT', data['options']['suppliers'])
        self.assertIn('骨干', data['options']['line_types'])

    def test_reject_all_empty_payload(self):
        response = self.client.post('/api/circuits', headers=self.headers, json={
            'supplier': '  ', 'supplier_circuit_id': '', 'circuit_id': '',
        })
        self.assertEqual(response.status_code, 400)

    def test_list_filters_and_keyword(self):
        self.create_circuit()
        self.create_circuit(
            supplier='CT', supplier_circuit_id='20-127306',
            circuit_id='MoscowCTVPN53810A', line_type='客户', line_status='待确认', remark='',
        )

        response = self.client.get('/api/circuits?supplier=CT', headers=self.headers)
        data = response.get_json()['data']
        self.assertEqual(data['total'], 1)
        self.assertEqual(data['rows'][0]['supplier'], 'CT')

        response = self.client.get('/api/circuits?line_type=骨干', headers=self.headers)
        self.assertEqual(response.get_json()['data']['total'], 1)

        # 关键字匹配 Circuit ID 中的片段（含特殊字符）
        response = self.client.get('/api/circuits?keyword=RT-STO/CTEUR', headers=self.headers)
        self.assertEqual(response.get_json()['data']['total'], 1)

        response = self.client.get('/api/circuits?keyword=MoscowCTVPN53810A', headers=self.headers)
        self.assertEqual(response.get_json()['data']['total'], 1)

    def test_update_and_delete(self):
        created = self.create_circuit()

        response = self.client.patch(
            f"/api/circuits/{created['id']}",
            headers=self.headers,
            json={'line_status': '已拆机', 'remark': '更新后的备注'},
        )
        self.assertEqual(response.status_code, 200)
        updated = response.get_json()['data']
        self.assertEqual(updated['line_status'], '已拆机')
        self.assertEqual(updated['remark'], '更新后的备注')
        self.assertEqual(updated['circuit_id'], MULTI_LINE_CIRCUIT_ID)

        response = self.client.delete(f"/api/circuits/{created['id']}", headers=self.headers)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.db.count_supplier_circuits(), 0)

        response = self.client.delete(f"/api/circuits/{created['id']}", headers=self.headers)
        self.assertEqual(response.status_code, 404)

    def test_import_preview_then_confirm_full_replace(self):
        self.create_circuit(supplier_circuit_id='will-be-replaced')

        import_bytes = build_import_workbook_bytes([
            {
                'supplier': 'RT',
                'supplier_circuit_id': '1233043(Stokholm - Beijing 10Gn1)\n20-267741',
                'circuit_id': 'HEIHE/CT-STO/CTEUR 64V4C001(OTU2)*TD\nHEIHE/CT-STO/CTEUR 64V4C001(OTU2)-new*TD',
                'line_type': '骨干',
                'line_status': '正常',
                'remark': '',
            },
            {
                'supplier': 'RT',
                'supplier_circuit_id': 751630,
                'circuit_id': '',
                'line_type': '语音电路',
                'line_status': '正常',
                'remark': '运营商邮件直接转发',
            },
        ])

        # 预览阶段：不写入数据库
        response = self.client.post(
            '/api/circuits/import',
            headers=self.headers,
            data={'file': (io.BytesIO(import_bytes), '线路表.xlsx')},
            content_type='multipart/form-data',
        )
        self.assertEqual(response.status_code, 200)
        preview = response.get_json()['data']
        self.assertEqual(preview['total'], 2)
        self.assertIn('\n', preview['rows'][0]['circuit_id'])
        self.assertEqual(preview['rows'][1]['supplier_circuit_id'], '751630')
        self.assertEqual(self.db.count_supplier_circuits(), 1)

        # 确认阶段：全量替换
        response = self.client.post(
            '/api/circuits/import',
            headers=self.headers,
            data={'file': (io.BytesIO(import_bytes), '线路表.xlsx'), 'confirm': 'true'},
            content_type='multipart/form-data',
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['data']['total'], 2)

        rows = self.db.get_all_supplier_circuits()
        self.assertEqual(len(rows), 2)
        self.assertNotIn('will-be-replaced', [row['supplier_circuit_id'] for row in rows])
        self.assertEqual(rows[0]['circuit_id'], 'HEIHE/CT-STO/CTEUR 64V4C001(OTU2)*TD\nHEIHE/CT-STO/CTEUR 64V4C001(OTU2)-new*TD')

    def test_import_rejects_unknown_headers(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['A', 'B', 'C'])
        ws.append(['1', '2', '3'])
        buffer = io.BytesIO()
        wb.save(buffer)

        response = self.client.post(
            '/api/circuits/import',
            headers=self.headers,
            data={'file': (io.BytesIO(buffer.getvalue()), 'bad.xlsx')},
            content_type='multipart/form-data',
        )
        self.assertEqual(response.status_code, 400)

    def test_export_roundtrip(self):
        self.create_circuit()
        response = self.client.get('/api/circuits/export')
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response.headers['Content-Type'])

        wb = load_workbook(io.BytesIO(response.data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        self.assertEqual(rows[0][:6], ('Supplier', 'Supplier Circuit ID', 'Circuit ID', '类型', '线路状态', '备注'))
        self.assertEqual(rows[1][2], MULTI_LINE_CIRCUIT_ID)

    def test_query_supplier_circuits_uses_database(self):
        self.create_circuit(supplier_circuit_id='1160515(Stokholm - Beijing; VC4S37 )')
        self.create_circuit(
            supplier='CT', supplier_circuit_id='999999',
            circuit_id='Other-Line/01', line_type='客户',
        )

        results = api_server.query_supplier_circuits('RT', ['1160515'])
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['supplier'], 'RT')
        self.assertEqual(results[0]['circuit_id'], MULTI_LINE_CIRCUIT_ID)
        self.assertEqual(results[0]['line_type'], '骨干')
        self.assertIn('中继线路', results[0]['remark'])

        # 供应商不匹配时返回空
        self.assertEqual(api_server.query_supplier_circuits('XX', ['1160515']), [])
        # 空关键字返回空
        self.assertEqual(api_server.query_supplier_circuits('RT', []), [])

    def test_migrate_from_xlsx(self):
        xlsx_path = Path(self.tmpdir.name) / '线路表.xlsx'
        xlsx_path.write_bytes(build_import_workbook_bytes([
            {
                'supplier': 'RT',
                'supplier_circuit_id': '20-127306',
                'circuit_id': 'MoscowCTVPN53810A',
                'line_type': '客户',
                'line_status': '正常',
                'remark': '',
            },
        ]))

        count = supplier_circuits.migrate_supplier_circuits_from_xlsx(db=self.db, xlsx_path=xlsx_path)
        self.assertEqual(count, 1)
        self.assertEqual(self.db.count_supplier_circuits(), 1)

        # 数据库已有数据时不重复迁移
        count = supplier_circuits.migrate_supplier_circuits_from_xlsx(db=self.db, xlsx_path=xlsx_path)
        self.assertEqual(count, 0)
        self.assertEqual(self.db.count_supplier_circuits(), 1)


if __name__ == '__main__':
    unittest.main()
